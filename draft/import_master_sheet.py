#!/usr/bin/env python3
"""Import the Fantasy Football Master Sheet — the league's founding document.

Parses every year tab (2016-2027), the Total Winnings tab (career W-L + per-year
money), Payments (2026 dues), and Votes into a structured archive with provenance
'source: master_sheet'. Cross-checks the 2023-2025 tabs against money_history
(they must reconcile — sheet and Sleeper harvest are independent sources). Career
totals derive from the YEAR columns, never the sheet's stale Total column.

Run: python draft/import_master_sheet.py
"""
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
ARCHIVE_XLSX = next((HERE / "archive" / "L2").glob("master-sheet.*.xlsx"), None)
OUT = HERE / "data" / "master_sheet_archive.json"

# Sheet uses first names + a couple of typos (Dyaln, Dave). Normalize to the
# identity table's real names.
NAME_FIX = {"Dyaln": "Dylan", "Dave": "David", "Michael Hagen": "Michael", "Hagen": "Michael"}
def fix_name(n):
    if n is None:
        return None
    n = str(n).strip()
    return NAME_FIX.get(n, n)


def _rows(ws):
    return [[c for c in r] for r in ws.iter_rows(values_only=True)]


def parse_year(ws):
    rows = _rows(ws)
    out = {"regular_season": {}, "playoffs": {}, "standings": [], "draft_order": [], "trades": []}
    # Header: "YYYY ($BUYIN)"
    hdr = str((rows[0][0] if rows and rows[0] else "") or "")
    m = re.match(r"(\d{4})\s*\(\$?([\d,]+)\)", hdr)
    if m:
        out["year"] = int(m.group(1))
        out["buy_in"] = int(m.group(2).replace(",", ""))
    # Trades can live in a single newline-delimited cell (2022: "TRADES\n*..\n*..").
    for r in rows:
        for c in r:
            if isinstance(c, str) and "trade" in c.lower() and "*" in c:
                for line in c.split("\n"):
                    line = line.strip()
                    if line.startswith("*"):
                        out["trades"].append(line.lstrip("*").strip())
    # Per-owner money table (2023+ tabs, cols J-O: name, Weekly, RS, Playoff, Total, Bets).
    header = [str(c or "").strip() for c in (rows[0] if rows else [])]
    if "Weekly Winnings" in header:
        wcol = header.index("Weekly Winnings")   # name is one col left of it
        ncol = wcol - 1
        pom = {}
        for r in rows[1:]:
            nm = fix_name(r[ncol]) if ncol < len(r) else None
            if not nm or nm.lower() == "total":
                continue
            def num(i):
                return r[i] if i < len(r) and isinstance(r[i], (int, float)) else 0
            pom[nm] = {"weekly": num(wcol), "regular_season": num(wcol + 1),
                       "playoffs": num(wcol + 2), "total": num(wcol + 3)}
        if pom:
            out["per_owner_money"] = pom
    # Section scan (column A labels).
    section = None
    for r in rows:
        a = str((r[0] if len(r) > 0 else "") or "").strip()
        b = r[1] if len(r) > 1 else None
        c = r[2] if len(r) > 2 else None
        d = r[3] if len(r) > 3 else None
        low = a.lower()
        if low.startswith("regular season standings"):
            section = "standings"; continue
        if low.startswith("draft order"):
            section = "draft_order"; continue
        if low == "regular season":
            section = "rs"; continue
        if low == "playoffs":
            section = "po"; continue
        if low == "trades":
            section = "trades"; continue
        rank = re.match(r"(\d+)(st|nd|rd|th)$", a)
        if section == "rs" and rank:
            out["regular_season"][a] = {"pct": b, "amount": c, "winner": fix_name(d)}
        elif section == "po" and rank:
            out["playoffs"][a] = {"pct": b, "amount": c, "winner": fix_name(d)}
        elif section == "standings" and rank:
            out["standings"].append({"rank": a, "owner": fix_name(b)})
        elif section == "draft_order" and rank:
            out["draft_order"].append({"rank": a, "owner": fix_name(b)})
        # 2022 lays standings in cols C/D alongside draft order in A/B.
        if section == "draft_order" and len(r) > 3:
            cc = str((r[2] if r[2] is not None else "")).strip()
            if re.match(r"\d+(st|nd|rd|th)$", cc):
                nm = fix_name(r[3])
                if nm and not any(s["rank"] == cc and s["owner"] == nm for s in out["standings"]):
                    out["standings"].append({"rank": cc, "owner": nm})
    # Nominal pot = buy_in x 10 owners. RS+PO sum is tracked separately; from 2023
    # the weekly-high ($1,500) is a SEPARATE pool not in the RS/PO rows, so the
    # RS+PO sum alone understates the pot — buy_in x 10 is the authoritative pot.
    amts = [v.get("amount") for v in list(out["regular_season"].values()) + list(out["playoffs"].values())
            if isinstance(v.get("amount"), (int, float))]
    out["rs_po_distributed"] = round(sum(amts), 2) if amts else None
    out["pot"] = (out.get("buy_in") or 0) * 10 or None
    return out


def parse_total_winnings(ws):
    rows = _rows(ws)
    header = [str(c or "").strip() for c in rows[0]]
    year_cols = {i: int(h) for i, h in enumerate(header) if re.match(r"^\d{4}$", h)}
    total_col = header.index("Total") if "Total" in header else None
    owners = {}
    for r in rows[1:]:
        name = fix_name(r[0])
        if not name or name.lower() == "total":
            continue
        by_year = {str(year_cols[i]): (r[i] if i < len(r) and isinstance(r[i], (int, float)) else 0)
                   for i in year_cols}
        career = round(sum(v for v in by_year.values()), 2)
        sheet_total = r[total_col] if (total_col is not None and total_col < len(r)
                                       and isinstance(r[total_col], (int, float))) else None
        owners[name] = {
            "wins": r[1], "loss": r[2], "tie": r[3], "win_pct": r[4],
            "by_year": by_year,
            "career_from_years": career,          # AUTHORITATIVE
            "sheet_total": sheet_total,           # may be STALE
            "stale": (sheet_total is not None and abs(sheet_total - career) > 0.01),
            "stale_delta": (round(career - sheet_total, 2) if sheet_total is not None else None),
        }
    return owners


def parse_payments(ws):
    rows = _rows(ws)
    out = {}
    for r in rows[1:]:
        name = fix_name(r[0])
        if not name or name.lower() == "total":
            continue
        # -400 = owes the full buy-in (unpaid); 0 or positive = paid/settled.
        val = r[1] if len(r) > 1 and isinstance(r[1], (int, float)) else None
        out[name] = {"raw_2026": val, "paid": (val is not None and val > -400)}
    return out


def main():
    if ARCHIVE_XLSX is None:
        print("ERROR: archived master sheet not found under draft/archive/L2/", file=sys.stderr)
        return 1
    import openpyxl
    import hashlib
    src_hash = hashlib.sha256(ARCHIVE_XLSX.read_bytes()).hexdigest()
    wb = openpyxl.load_workbook(ARCHIVE_XLSX, data_only=True)

    seasons = {}
    for name in wb.sheetnames:
        if re.match(r"^\d{4}$", name):
            seasons[name] = parse_year(wb[name])
    total_winnings = parse_total_winnings(wb["Total Winnings"]) if "Total Winnings" in wb.sheetnames else {}
    payments = parse_payments(wb["Payments"]) if "Payments" in wb.sheetnames else {}
    votes = [str(r[1]).strip() for r in _rows(wb["Votes"]) if len(r) > 1 and r[1]] if "Votes" in wb.sheetnames else []

    archive = {
        "provenance": "source: master_sheet",
        "source_file": ARCHIVE_XLSX.name,
        "source_sha256": src_hash,
        "note": "The league's founding document — Est. 2016. Pre-Sleeper seasons (2016-2022) exist NOWHERE else. Career money derives from by_year, never sheet_total (which is stale for owners with 2025 winnings).",
        "seasons": seasons,
        "total_winnings": total_winnings,
        "payments_2026": payments,
        "votes_pending": votes,
    }
    OUT.write_text(json.dumps(archive, indent=2))

    # --- Report: import counts per year ---
    print("MASTER SHEET IMPORT — counts per year (source: master_sheet)")
    for y in sorted(seasons):
        s = seasons[y]
        print(f"  {y}: buy_in=${s.get('buy_in')} pot=${s.get('pot')} "
              f"RS={len(s['regular_season'])} PO={len(s['playoffs'])} "
              f"standings={len(s['standings'])} draft_order={len(s['draft_order'])} trades={len(s['trades'])}")
    print(f"  Total Winnings: {len(total_winnings)} owners (W-L + per-year money)")
    print(f"  Payments (2026 dues): {len(payments)} owners")
    print(f"  Votes pending: {len(votes)}")

    # --- CROSS-CHECK 2023-2025 sheet money vs money_history (independent sources) ---
    print("\nCROSS-CHECK — master sheet per-owner money vs money_history (must reconcile):")
    try:
        sys.path.insert(0, str(HERE))
        import money_history as MH
        mh = MH.analyse()
        # money_history career totals keyed by owner_id; map to real names via identity.
        idmap = json.loads((HERE / "config" / "identity_map.json").read_text())
        id_to_name = {}
        for real, rec in (idmap.get("by_real_name") or {}).items():
            if rec.get("owner_id"):
                id_to_name[str(rec["owner_id"])] = real
        # money_history 3-season (2023-25) career total per resolved owner.
        mh_career = {}
        for row in mh.get("dollar_standings", []):
            nm = id_to_name.get(str(row["name"]))
            if nm:
                mh_career[nm] = round(row["total_$"], 2)
        # sheet 2023-25 career from by_year.
        agree = 0; flags = []
        for real, d in total_winnings.items():
            sheet_3yr = round(sum(d["by_year"].get(str(y), 0) for y in (2023, 2024, 2025)), 2)
            fn = {"Michael": "Michael Hagen"}.get(real, real)
            mhv = mh_career.get(real, mh_career.get(fn))
            if mhv is not None:
                if abs(sheet_3yr - mhv) < 0.01:
                    agree += 1
                else:
                    flags.append(f"{real}: sheet ${sheet_3yr} vs money_history ${mhv} (Δ ${round(sheet_3yr-mhv,2)})")
        print(f"  resolved owners agreeing: {agree}; flags: {len(flags)}")
        for f in flags:
            print(f"    🚩 {f}")
        if not flags and agree:
            print("  ✅ every resolved owner's 2023-25 money reconciles — sheet and Sleeper harvest agree.")
    except Exception as e:
        print(f"  (cross-check skipped: {e})")

    # --- Staleness log (the documented data-spine example) ---
    stale = {n: d["stale_delta"] for n, d in total_winnings.items() if d["stale"]}
    print("\nSTALE Total-column owners (career_from_years − sheet_total):")
    for n, delta in sorted(stale.items()):
        print(f"  {n}: +{delta} excluded from the sheet Total (career must use year columns)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
